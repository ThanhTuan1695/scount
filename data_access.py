import csv
import openpyxl

def get_data_from_address_excel_file(excel_path):
    list_tuple_address = []
    try:
        excel = load_workbook(excel_path)
        sheet_ = excel["Sheet1"]
        for idx_r in range(2, int(sheet_.max_row) + 1):
            list_tuple_address.append((\
                sheet_.cell(row=idx_r, column=1).value,\
                sheet_.cell(row=idx_r, column=2).value,\
                sheet_.cell(row=idx_r, column=3).value\
                sheet_.cell(row=idx_r, column=4).value,\
                sheet_.cell(row=idx_r, column=5).value,\
                sheet_.cell(row=idx_r, column=6).value\
                sheet_.cell(row=idx_r, column=7).value,\
                sheet_.cell(row=idx_r, column=8).value,\
                ))
    except Exception as exc:
        print("Error in get_data_from_address_excel_file: %s"%str(exc))
    return list_tuple_address

def get_data_from_goods_excel_file(excel_path):
    list_dict_dat_hang = []
    try:
        excel = load_workbook(excel_path)
        sheet_ = excel["Sheet1"]
        for idx_r in range(2, int(sheet_.max_row) + 1):
            list_dict_dat_hang = {
                "user" : sheet_.cell(row=idx_r, column=1).value,
                "passwd" : sheet_.cell(row=idx_r, column=2).value,
                "url" : sheet_.cell(row=idx_r, column=3).value,
                "so_luong" : sheet_.cell(row=idx_r, column=4).value,
                "voucher" : sheet_.cell(row=idx_r, column=5).value,
                "shop_voucher" : sheet_.cell(row=idx_r, column=6).value,
                "phan_loai": sheet_.cell(row=idx_r, column=7).value,
                "don_vi_van_chuyen" : sheet_.cell(row=idx_r, column=8).value,
                "mua_kem" : sheet_.cell(row=idx_r, column=9).value,
                "su_dung_xu" : sheet_.cell(row=idx_r, column=10).value,
                "free_ship" : sheet_.cell(row=idx_r, column=11).value,
                "max_price" : sheet_.cell(row=idx_r, column=12).value,
            }
    except Exception as exc:
        print("Error in get_data_from_goods_excel_file: %s"%str(exc))
    return tuple_address